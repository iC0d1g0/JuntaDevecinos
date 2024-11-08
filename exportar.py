import os

from openpyxl import Workbook
from tkinter import messagebox
from crud import Crud
"""
Este objeto llama  Exel, recibe un parametro de nombre tabla y exporta esa tabla a un 
documento exel o xlsx. 
Maneja la execion logica de la existencia de algun otro archivo con el mismo nombre y no lo reemplaza, 
si no que crea uno con el mismo nombre y agrega 1++.

"""
""""
TODOS ESTE CODIGO DEBE SER MODIFICADO PARA QUE NO ACCEDA A LA BASE DE DATOS PARA LEER. 
1. UN UNICO OBJECTO QUE INTERATUE CON LA BASE DE DATO CRUD. 
2.EL OBJECTO EXPORTAR : LLAMA EL OBJECTO QUE INTERACTUA, LE PASE EL NOMBRE DE LA TABLA
 Y LAS COLUMNAS QUE QUIERE ACCEDER Y LE PIDE QUE LE DEBUELVA LA INFO. 
3. LA INFORMACION DEBUELTA, DEBE SER CONVERTIDA A XLSM Y PARA ESO ENTRA EN ACCION EL OBJECTO Exel.

"""
class Excel(Crud): #FUnciona!!!!
         # Conectar a la base de datos
        """
        Llamar el objecto conexion
        """
        def __init__(self,tabla):
            super().__init__()
            self.cursor=self.conexion.cursor()
            self.tablapropia=tabla
            """self.datos=self.cru.leer()
            self.description=self.cru.exportar()"""
            self.archivo_xlsx = f"C:/Users/{os.getenv('USERNAME')}/Desktop/Reporte_JV"
            self.exportar_db_a_xlsx()

        def exportar_db_a_xlsx(self):
            self.cursor.execute(f"SELECT * FROM {self.tablapropia}")
            self.datos=self.cursor.fetchall()
            # Obtener los nombres de las columnas
            nombres_columnas = [descripcion[0] for descripcion in self.cursor.description]
            # Quiero utilizar esto para leer las columnas de las base de datos
            
            # Crear un nuevo libro de trabajo (Workbook)
            libro_trabajo = Workbook() #
            hoja = libro_trabajo.active # Crear una hoja en el libro de trabajo
            

            """
            lista_tupla=[(1, elemento), (2, elemnto), (3, elemento)]
            columna_idx=[1,2,3]
            columna_nombre=[elemento,elemento,elemento]
            """
            # Escribir los nombres de las columnas en la primera fila
            for columna_idx, columna_nombre in enumerate(nombres_columnas, start=1):
                hoja.cell(row=1, column=columna_idx, value=columna_nombre)#row=1Esto significa que ocupara la primera fila
                #de la hoja en excell. 
                #significa que ocupara el indice de la coluna (Ejemplo: primero sera en la columna A, luego columna B...)


            # Escribir los datos en las filas siguientes
            for fila_idx, fila_datos in enumerate(self.datos, start=2):
                for columna_idx, valor in enumerate(fila_datos, start=1):
                    hoja.cell(row=fila_idx, column=columna_idx, value=valor)
                    
            # Guardar el libro de trabajo en un archivo .xlsx
            
            # Verificar si existe el archivo con el mismo nombre: 
            compara=os.path.exists('{}.xlsx'.format(self.archivo_xlsx))
            if compara: #comparamos el nombre del archivo en la misma direccion. 
                pregunta=messagebox.askyesno('Archivo existe!!', 'Deseas crear otro archivo?')
                if pregunta:
                    contador=1
                    self.archivo_xlsx+='_'
                    #Este ciclo solo sirve para confirmar los archivos existentes con el mismo nombre. 
                    while  os.path.exists('{}{}.xlsx'.format(self.archivo_xlsx,contador)): 
                        
                        contador+=1
                    libro_trabajo.save('{}{}.xlsx'.format(self.archivo_xlsx,contador))
                    messagebox.showinfo("Exportacion", 'Guardado en: {}{}.xlsx'.format(self.archivo_xlsx,contador))
                else: 
                    messagebox.showinfo('Archivo', 'No se ha guardado ni reemplazado el archivo') 
               
            else:
                libro_trabajo.save('{}.xlsx'.format(self.archivo_xlsx))
                messagebox.showinfo("Exportacion", "Datos Exportados en ruta: {}".format(self.archivo_xlsx))
            # Cerrar la conexión
            self.cursor.close()
            

     

